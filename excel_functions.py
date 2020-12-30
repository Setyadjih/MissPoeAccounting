import openpyxl

from logging import getLogger

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side

from app import ExcelItem

# Today as 30-Mar-19
DATE_FORMAT = "dd-mmm-yy"

COMMA_FORMAT = "#,##0"
RP_FORMAT = u'_("Rp"* #,##0_);_("Rp"* (#,##0);_("Rp"* "-"_);_(@_)'

cat_sheets = {"LIST", "ITEM LIST", "Fresh", "Sundries", "Packaging",
              "Utensils", "Appliances", "Cleaning"}


def init_catsheet(file, categories: dict, logger):
    logger = logger if logger else getLogger()
    # Load excel file path
    # Due to openpyxl's structure, we need the data_only=False wb to save
    # formula
    input_wb = openpyxl.load_workbook(file, data_only=False)

    # Clear out all category sheets, leaving the header only
    for category in categories["CATEGORIES"]:
        category_sheet = input_wb[category]
        max_row = category_sheet.max_row
        category_sheet.delete_rows(3, max_row)

    # default dict
    done_set = {"None", " "}
    skip_list = categories["CATEGORIES"] + categories["MISC"]

    # Iterate over all vendor sheets
    vendor_sheets = [_ for _ in input_wb.sheetnames if _ not in skip_list]
    logger.debug(f"VENDORS: {vendor_sheets}")
    for sheet_name in vendor_sheets:
        sheet: Worksheet = input_wb[sheet_name]
        logger.info(f"Sheet: {sheet}")
        items = next(sheet.iter_cols(min_col=2, max_col=2, min_row=3, values_only=True))
        for item in items:
            # Early check done_list to skip checking in cat sheet
            if not item or item in done_set:
                logger.debug("Skipping item")
                continue

            row = items.index(item) + 3
            logger.debug(f"ROW: {row}, item: {item}")

            # Try to get data from row. If missing data, give defaults
            try:
                isi_unit = sheet[f"I{row}"].value
                if not isi_unit:
                    isi_unit = sheet[f"E{row}"].value

            except IndexError:
                logger.error("Isi error, assigning g")
                isi_unit = "g"

            # Check for common typos in categories
            try:
                categ: str = sheet[f"K{row}"].value
                if not categ:
                    logger.debug("Assigning to default")
                    categ = "Fresh"

                corrections = {
                    "Utensil": "Utensils",
                    "Packing": "Packaging",
                    "Sundry": "Sundries",
                    "Alat tulis": "Stationary",
                    "Stationery": "Stationary",
                    "Appliance": "Appliances"
                }

                category_check = categ.strip().lower().capitalize()
                if category_check in corrections.keys():
                    categ = corrections[category_check]

            except IndexError:
                logger.error("category error, assigning Fresh")
                categ = "Fresh"

            logger.debug(f"Category: {categ}")

            # Check if item is already in cat sheet
            cat_items = next(input_wb[categ].iter_cols(1, 1, values_only=True))
            if item in cat_items:
                logger.info("Item already in category, skipping")
                continue

            logger.info(f"Appending {item} to {categ}")
            excel_item = ExcelItem(
                name=item, vendor=sheet_name, isi_unit=isi_unit, category=categ
            )
            update_cat(skip_list, file, excel_item, input_wb, logger)
            done_set.add(item)
    logger.info("All done with init")


def write_to_excel(skips, date, file, excel_item, logger=None):
    """ Write the given data to the purchasing excel sheet

    :param skips sheets to skip
    :type skips: List[str]
    :param str date: date of purchase
    :param str file: file path to excel sheet to edit
    :param excel_item: ExcelItem with data
    :type excel_item: ExcelItem
    :param logger: logger pass through
    """
    logger = logger if logger else getLogger()

    # Load excel file path
    # Need the data_only=False wb to save formula
    input_wb = openpyxl.load_workbook(file, data_only=False)

    # Create vendor sheet if new
    input_vendor = input_wb[excel_item.vendor]
    last_row = input_vendor.max_row + 1

    # iterate backwards until last_row is after a row with data
    while not input_vendor[f"B{last_row-1}"].value:
        last_row -= 1

    logger.debug(f"max row = {input_vendor.max_row}, last row = {last_row}")
    logger.debug(f"Appending item data")
    input_vendor[f"A{last_row}"] = date
    input_vendor[f"B{last_row}"] = excel_item.name
    input_vendor[f"C{last_row}"] = excel_item.brand
    input_vendor[f"D{last_row}"] = excel_item.quantity
    input_vendor[f"E{last_row}"] = excel_item.unit
    input_vendor[f"F{last_row}"] = excel_item.cost
    input_vendor[f"G{last_row}"] = f'=D{last_row}*F{last_row}'
    input_vendor[f"H{last_row}"] = excel_item.isi
    input_vendor[f"I{last_row}"] = excel_item.isi_unit
    input_vendor[f"J{last_row}"] = f"=G{last_row}/H{last_row}"
    input_vendor[f"K{last_row}"] = excel_item.category

    # format cells for Rupiah
    logger.debug("Assigning format strings")
    date_cell = input_vendor.cell(last_row, 1)
    date_cell.number_format = DATE_FORMAT

    cost_cell = input_vendor.cell(last_row, 6)
    cost_cell.number_format = RP_FORMAT

    total_cell = input_vendor.cell(last_row, 7)
    total_cell.number_format = RP_FORMAT

    isi_cell = input_vendor.cell(last_row, 8)
    isi_cell.number_format = COMMA_FORMAT

    per_unit_cell = input_vendor.cell(last_row, 10)
    per_unit_cell.number_format = RP_FORMAT

    update_cat(skips, file, excel_item, input_wb, logger)


def update_cat(skips, file, excel_item, input_wb, logger):
    """ Update item entry in category sheet

    :param skips: sheets to skip
    :param str file: file path
    :param excel_item: ExcelItem for item data
    :param input_wb: Workbook to save and input form
    :param logger: logger passthrough
    """
    # Append to costing according to category
    category = excel_item.category

    if category not in input_wb.sheetnames:
        logger.debug(f"{category} not found, creating")
        input_wb.create_sheet(category)
        cat_sheet: Worksheet = input_wb[category]
        cat_sheet['A1'] = category.upper()
        cat_sheet['A2'] = 'MATERIAL'
        cat_sheet['B2'] = 'UNIT'
        cat_sheet['C2'] = 'PRICE'

        # Apply style to docs
        for i in range(1, 4):
            cell = cat_sheet.cell(row=2, column=i)
            cell.border = Border(bottom=Side(style='thick'))

    input_wb.save(file)
    logger.debug(f"Assigning {excel_item.name} to {category}")

    # Handle cat sheet updating
    update_avg_formula(excel_item, input_wb, skips, logger)

    input_wb.save(filename=file)


def update_avg_formula(excel_item, workbook, skips, logger=None):
    """Calculate average price for each item, total quantity, total units
    Average formula template:
    SUM(
        SUMIF('[VENDORSHEET]'!B:B, A[ROW], '[VENDORSHEET]'!J:J),
        SUMIF('[VENDERSHEET]'!....
    )
    /
    SUM(
        COUNTIF('[VENDORSHEET]'!B:B, A[ROW]),
        COUNTIF('[VENDORSHEET]'!....
    )
    :param excel_item: ExcelItem with data
    :param workbook: workbook to read from
    :param skips: sheets to avoid
    :param logger: logger pass through
    :return average
    """
    # Check for item in each ws
    # for each ws with item, add SUMIF to final formula
    logger = logger if logger else getLogger()

    category = excel_item.category
    vendor_check = excel_item.vendor

    # check if item in list
    cat_items = next(workbook[category].iter_cols(1, 1, values_only=True))
    if excel_item.name in cat_items:
        logger.debug("Item exists, checking vendor in formula")

        # compensate for 1 based index
        row = cat_items.index(excel_item.name) + 1
        avg_formula: str = workbook[category][f"C{row}"].value

        if vendor_check in avg_formula:
            logger.debug("Vendor in formula, returning")
            return
        else:
            logger.debug("Vendor is new, adding to formula")
            sumif_str = f"SUMIF('{vendor_check}'!B:B, A{row}, '{vendor_check}'!J:J),"
            countif_str = f"COUNTIF('{vendor_check}'!B:B, A{row}),"

            # Insert at divisor
            # add count to divisor sum
            # Rewrite formula back to wb
            avg_formula = avg_formula.replace(") /", f"{sumif_str}) /")
            avg_formula = avg_formula[:-1] + countif_str + ")"
            workbook[category][f"C{row}"] = avg_formula

    else:
        logger.info(f"Item is new, creating {category} entry")
        workbook[category].append({'A': excel_item.name, 'B': excel_item.isi_unit})
        row = workbook[category].max_row

        # Iterating through each worksheet to find all vendors with item
        logger.debug(f"Beginning iteration")
        price_count = ""
        entry_count = ""

        try:
            vendors = [_ for _ in workbook.sheetnames if _ not in skips]
            # Check each vendor for item
            for vendor in vendors:
                items = next(workbook[vendor].iter_cols(2, 2, values_only=True))
                if excel_item.name in items:
                    price_count += f"SUMIF('{vendor}'!B:B, A{row}, '{vendor}'!J:J),"
                    entry_count += f"COUNTIF('{vendor}'!B:B, A{row}),"

            # Combine average formula and apply to workbook
            avg_formula = f"=SUM({price_count})/SUM({entry_count})"
            workbook[category][f"C{row}"] = avg_formula

            price_cell_obj = workbook[category].cell(row=row, column=3)
            price_cell_obj.number_format = RP_FORMAT
        except Exception as e:
            logger.error(f"ERROR: {e}")
