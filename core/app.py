import sys
import time
from datetime import datetime
import shutil
from pathlib import Path
import subprocess

from PySide6.QtWidgets import (
    QApplication,
    QWidget,
    QTableWidgetItem,
    QFileDialog,
    QMessageBox,
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QAction
from openpyxl import load_workbook
import pyautogui

from core.utils import (
    init_logger,
    get_file_handler,
    write_default_categories_file,
    read_categories_file,
    get_skip_list,
)
from resources.pembelian_ui_ss import Ui_pembelian
from core.excel_functions import write_to_excel, init_catsheet, import_records
from core.constants import APP_VERSION, DATE, CAT_REF, ExcelItem, LOGGER_NAME, Status


# noinspection SpellCheckingInspection
class PembelianWidget(QWidget):
    def __init__(self, test=False, parent=None):
        super(PembelianWidget, self).__init__(parent)
        self.ui = Ui_pembelian()
        self.ui.setupUi(self)
        self.ui.date_line.setText(DATE)
        self.logger = init_logger(LOGGER_NAME)
        self.logger.info("Initializing program")

        self.cat_items_dict: dict[str, list[ExcelItem]] = {}

        # Context menu setup
        self.ui.commit_table.setContextMenuPolicy(Qt.ActionsContextMenu)
        self.del_row_action = QAction(self, text="Delete Table Row")

        # noinspection PyUnresolvedReferences
        self.del_row_action.triggered.connect(self.delete_table_row)
        self.ui.commit_table.addAction(self.del_row_action)

        self.setWindowTitle(f"Poe Excel Automator {APP_VERSION}")
        self.ui.status_bar.setText("Ready for input.")
        self.ui.item_line.hide()

        # Default test button to hide
        self.ui.test_button.hide()
        if test:
            self.ui.test_button.show()
            self.ui.test_button.clicked.connect(self.test_func)

        # Setup Category dictionary and file
        self.logger.debug(f"Checking for CAT_REF at: {Path(CAT_REF).absolute()}")
        if not Path(CAT_REF).exists():
            self.logger.info("Creating new cat ref file")
            write_default_categories_file()
        self.categories = read_categories_file()

        self.skip_list = get_skip_list()
        self.ui.category_combo.clear()
        self.ui.category_combo.addItems(self.categories["CATEGORIES"])

        # Setup button Tooltips
        self.ui.file_browse_button.setToolTip("Select workbook to make active")
        self.ui.add_vendor_button.setToolTip("Add entry")
        self.ui.import_button.setToolTip("Import price data from previous to active workbook")
        self.ui.init_button.setToolTip("Clear and recheck category items")
        self.ui.confirm_button.setToolTip("Confirm entries to excel")

        # Hookup buttons
        self.ui.file_browse_button.clicked.connect(self.get_excel_sheet)
        self.ui.new_item_check.stateChanged.connect(self.item_input_toggle)
        self.ui.item_combo.currentIndexChanged.connect(self.item_unit_lock)
        self.ui.add_vendor_button.clicked.connect(self.add_to_table)
        self.ui.confirm_button.clicked.connect(self.confirm_table)
        self.ui.init_button.clicked.connect(self.init_cat_button)
        self.ui.import_button.clicked.connect(self.import_data)
        self.ui.category_combo.currentIndexChanged.connect(self.load_cat_items)

    def load_cat_items(self):
        """Load all items in cateogory"""
        self.ui.item_combo.clear()
        current_cat = self.ui.category_combo.currentText()
        # incase invalid text
        if not current_cat:
            self.logger.error("Category is empty, could not load items")
            return

        for item in self.cat_items_dict[current_cat]:
            self.ui.item_combo.addItem(item.name, userData=item)

    def item_input_toggle(self):
        """Toggle item input style"""
        if self.ui.new_item_check.isChecked():
            self.ui.item_combo.hide()
            self.ui.item_line.show()
            self.ui.unit_combo.setEnabled(True)
            self.ui.isi_unit_combo.setEnabled(True)
        else:
            self.ui.item_combo.show()
            self.ui.item_line.hide()
            self.ui.item_line.clear()

    def item_unit_lock(self):
        """Lock item units to preexisting data"""
        if not self.ui.new_item_check.isChecked():
            item: ExcelItem = self.ui.item_combo.currentData()
            if not item:
                return
            if item.unit_isi:
                self.ui.isi_unit_combo.setCurrentText(item.unit_isi)
                self.ui.isi_unit_combo.setDisabled(True)
            else:
                self.ui.isi_unit_combo.setEnabled(True)

    def test_func(self):
        """Clear out category sheets"""
        input_wb = load_workbook(self.ui.xls_file_browser.text(), data_only=False)
        for cat in self.categories["CATEGORIES"]:
            category_sheet = input_wb[cat]
            max_row = category_sheet.max_row
            category_sheet.delete_rows(3, max_row)
        input_wb.save(self.ui.xls_file_browser.text())

    def make_backup(self):
        # Create a backup copy just in case
        self.logger.info("Saving backup.")
        file = self.ui.xls_file_browser.text()
        dest = Path(file).with_suffix(".bak")
        shutil.copyfile(file, dest)

    def safe_gui(self, title, func, *args):
        if pyautogui.getActiveWindowTitle() != title:
            self.logger.debug(f"Active Window: {pyautogui.getActiveWindowTitle()} does not match {title}")
            raise AssertionError(f"Bad Active Window: {pyautogui.getActiveWindowTitle()} != {title}")
        func(*args)
        time.sleep(1)

    def init_cat_button(self):
        """Start initialization of category sheets"""
        file = self.ui.xls_file_browser.text()
        if not file:
            self.__set_info("No file to write to!", Status.FAIL)
            return

        result = QMessageBox.warning(
            self,
            "Are you sure?",
            "Initializing the category data can take a long time, are you sure"
            " you want to do this?\n The category sheets will be cleared out.",
            QMessageBox.Cancel,
            QMessageBox.Ok,
        )
        if result != QMessageBox.Ok:
            return

        self.__set_info("Working on data....")
        self.make_backup()
        try:
            init_catsheet(file, self.categories)
        except Exception as e:
            self.logger.error(e)
            self.__set_info(f"Failed to init data! Error: {e}", Status.FAIL)
            return
        self.logger.info("Finished init")
        self.__set_info("All done!", Status.DONE)
        # subprocess.run(["start", "excel", file], shell=True)
        # time.sleep(2)
        # result = QMessageBox.information(self, "Loading File..", "Please Hit OK when file is loaded, then do not touch anything")
        # if result != QMessageBox.Ok:
        #     self.__set_info("Canceled auto cleaning.")
        #     return

        # self.__set_info("Start cleaning!")
        # self.logger.info("Start cleaning!")
        # self.logger.debug(f"finding window with title: {file.title().split('/')[-1]} - Excel")
        # pyautogui.getWindowsWithTitle(f"{file.title().split('/')[-1]} - Excel")[0].activate()
        # time.sleep(1)
        # title = pyautogui.getActiveWindowTitle()
        #
        # # Hacky as hell gui automation to fix excel formulas
        # for cat in self.categories["CATEGORIES"]:
        #     try:
        #         self.logger.info(f"Cleaning {cat} Max!")
        #         self.safe_gui(title, pyautogui.press, "f5")
        #         self.safe_gui("Go To", pyautogui.write, f"{cat}!E3")
        #         self.safe_gui("Go To", pyautogui.press, "enter")
        #         pyautogui.hotkey("ctrl", "shift", "end", interval=0.5)
        #         time.sleep(1)
        #         self.safe_gui(title, pyautogui.press, "f2")
        #         self.safe_gui(title, pyautogui.press, "enter")
        #         time.sleep(2)
        #         self.safe_gui(title, pyautogui.hotkey, "ctrl", "d")
        #         time.sleep(3)
        #     except AssertionError:
        #         QMessageBox.information(self, "Safety Cancel",
        #                                 "Excel is no longer the active window! Exiting just in case.")
        #         self.__set_info("Cleaning safety cancel triggered")
        #
        #         return
        # QMessageBox.information(self, "Finished!", "Finished Cleaning all categories!")
        self.__set_info("All done!", Status.DONE)

    def import_data(self):
        """Import data from previous workbook to current active workbook"""
        new_workbook = self.ui.xls_file_browser.text()
        if not new_workbook:
            self.__set_info("Please select Workbook to import to!", Status.FAIL)

        try:
            old_workbook = QFileDialog.getOpenFileName(filter="Old Workbook (*.xlsx)")[0]
        except KeyError as error:
            self.__set_info(f"Failed to pick sheet! Vendor doesn't exist.", Status.FAIL)
            self.logger.error(error)
            return
        except Exception as error:
            self.__set_info(f"Failed to pick sheet! Reason: {error}", Status.FAIL)
            self.logger.error(error)
            return
        self.__set_info("Transferring records...")
        import_records(old_workbook, new_workbook, self.categories)
        self.__set_info("Done Transferring!", Status.DONE)

    def delete_table_row(self):
        current_row = self.ui.commit_table.currentRow()
        self.ui.commit_table.removeRow(current_row)

    def get_excel_sheet(self):
        """Load Purchase Excelsheet and get vendors"""
        try:
            file_dir = QFileDialog.getOpenFileName(filter="Excel sheets (*.xlsx)")[0]
        except KeyError as error:
            self.__set_info(f"Failed to pick sheet! Vendor doesn't exist.", Status.FAIL)
            self.logger.error(error)
            return
        except Exception as error:
            self.__set_info(f"Failed to pick sheet! Reason: {error}", Status.FAIL)
            self.logger.error(error)
            return

        # File check
        self.ui.xls_file_browser.setText(file_dir)
        if not self.ui.xls_file_browser.text():
            self.__set_info("Did not get file path", Status.FAIL)
            return

        # Populate vendor drop down
        purchase_book = load_workbook(file_dir)
        skip_list = self.categories["CATEGORIES"] + self.categories["MISC"]
        vendor_sheets = [_ for _ in purchase_book.sheetnames if _ not in skip_list]
        for vendor in vendor_sheets:
            self.ui.vendor_combo.addItem(vendor)

        # populate category selection with item lists
        bad_cats = []
        for category in self.categories["CATEGORIES"]:
            cat_items = []
            try:
                for row in purchase_book[category].iter_rows(min_row=3, values_only=True):
                    name = row[0].strip()
                    # In case of missing item names or empty rows, skip
                    if not name:
                        continue

                    # Guard against missing units
                    unit_beli = row[1] if row[1] else "NA"
                    unit_isi = row[2] if row[2] else "NA"

                    cat_items.append(ExcelItem(name=name, unit_beli=unit_beli, unit_isi=unit_isi))
            except KeyError:
                self.logger.info(f"{category} not in Workbook")
                bad_cat_index = self.ui.category_combo.findText(category)
                bad_cats.append(bad_cat_index)

            self.cat_items_dict[category] = sorted(cat_items, key=lambda item: item.name)

        # Remove invalid categories from loaded sheet
        for cat in reversed(sorted(bad_cats)):
            self.ui.category_combo.removeItem(cat)

        # initial category population
        self.load_cat_items()

        # Add easy to access log with username
        log_dir = Path(file_dir).parent.joinpath("_LOG").as_posix()
        self.logger.addHandler(get_file_handler("excel_automator", log_dir))
        self.logger.debug("Init user logging")

    def clear_inputs(self):
        """Clear out input fields"""
        self.ui.vendor_combo.clear()
        self.ui.item_line.clear()
        self.ui.qty_spin.clear()
        self.ui.harga_spin.clear()
        self.ui.isi_spin.clear()
        self.__set_info("Cleared inputs!", status=Status.DONE)

    def find_existing_item_category(self, item):
        for category in self.cat_items_dict.keys():
            sanitized_items = [x.name for x in self.cat_items_dict[category]]
            if item.strip().lower() in sanitized_items:
                return category

    def add_to_table(self):
        # Table entry validation
        if (
            self.ui.harga_spin.value() == 0
            or self.ui.isi_spin.value() == 0
            or self.ui.qty_spin.value() == 0
            or not self.ui.date_line.text()
            or not self.ui.isi_unit_combo.currentText()
            or not self.ui.vendor_combo.currentText()
        ):
            self.__set_info("Values cannot be 0!", Status.FAIL)
            return

        # Check if item already exists in any category
        if self.ui.new_item_check.isChecked():
            sanitized_new_item = self.ui.item_line.text().strip().lower()
            sanitized_items = [item.name.strip().lower() for item_list in self.cat_items_dict.values() for item in item_list]

            if sanitized_new_item in sanitized_items:
                self.logger.debug(f"Found pre-existing item {self.ui.item_line.text()}")
                # Switch to category and select item if exists
                self.ui.new_item_check.setChecked(False)

                item_category = self.find_existing_item_category(sanitized_new_item)
                self.ui.category_combo.setCurrentIndex(self.ui.category_combo.findText(item_category))
                self.logger.debug(f"Found item in {item_category}")

                category_items = [x.name.strip().lower() for x in self.cat_items_dict[item_category]]
                item_index = category_items.index(sanitized_new_item)
                self.logger.debug(f"Found item index: {item_index}")
                self.ui.item_combo.setCurrentIndex(item_index)

        # Commit input to table
        row_count = self.ui.commit_table.rowCount()
        self.ui.commit_table.insertRow(row_count)
        new_row = row_count

        details = self.get_ui_details()
        for column, item in enumerate(details):
            self.ui.commit_table.setItem(new_row, column, item)
        self.__set_info("Added item to table")

    def get_ui_details(self) -> tuple:
        """Get all information from data fields returned as a set"""
        date_text = self.ui.date_line.text()
        date = datetime.strptime(date_text, "%d-%b-%y")
        total_cost = self.ui.qty_spin.value() * self.ui.harga_spin.value()
        unit_cost = total_cost / self.ui.isi_spin.value()
        date_data = QTableWidgetItem(self.ui.date_line.text())
        date_data.setData(Qt.UserRole, date)
        vendor_data = QTableWidgetItem(self.ui.vendor_combo.currentText())
        vendor_data.setData(Qt.UserRole, self.ui.vendor_combo.currentText())
        if not self.ui.merek_line.text():
            self.ui.merek_line.setText("")
        merek_data = QTableWidgetItem(self.ui.merek_line.text())
        merek_data.setData(Qt.UserRole, self.ui.merek_line.text())
        item_text = self.ui.item_combo.currentText()
        if self.ui.new_item_check.isChecked():
            item_text = self.ui.item_line.text().strip()
        item_data = QTableWidgetItem(item_text)
        item_data.setData(Qt.UserRole, item_text)
        qty_data = QTableWidgetItem(self.ui.qty_spin.text())
        qty_data.setData(Qt.UserRole, self.ui.qty_spin.value())
        unit_data = QTableWidgetItem(self.ui.unit_combo.currentText())
        unit_data.setData(Qt.UserRole, self.ui.unit_combo.currentText())
        harga_data = QTableWidgetItem(self.ui.harga_spin.text())
        harga_data.setData(Qt.UserRole, self.ui.harga_spin.value())
        total_data = QTableWidgetItem(str(total_cost))
        total_data.setData(Qt.UserRole, total_cost)
        isi_data = QTableWidgetItem(self.ui.isi_spin.text())
        isi_data.setData(Qt.UserRole, self.ui.isi_spin.value())
        isi_unit_data = QTableWidgetItem(self.ui.isi_unit_combo.currentText())
        isi_unit_data.setData(Qt.UserRole, self.ui.isi_unit_combo.currentText())
        unit_harga_data = QTableWidgetItem(str(unit_cost))
        unit_harga_data.setData(Qt.UserRole, unit_cost)
        category_data = QTableWidgetItem(self.ui.category_combo.currentText())
        category_data.setData(Qt.UserRole, self.ui.category_combo.currentText())
        details = (
            date_data,
            item_data,
            vendor_data,
            merek_data,
            qty_data,
            unit_data,
            harga_data,
            total_data,
            isi_data,
            isi_unit_data,
            unit_harga_data,
            category_data,
        )
        return details

    def confirm_table(self):
        """Commit table to Excel file"""
        self.logger.info("Executing table to excel file")

        file = self.ui.xls_file_browser.text()
        if not file:
            self.__set_info("No file to write to!", Status.FAIL)
            return

        if self.ui.commit_table.rowCount() == 0:
            self.__set_info("Nothing to write")
            return

        self.make_backup()
        for row in range(self.ui.commit_table.rowCount()):
            try:
                # Get values from item ranges as an ExcelItem
                date = self.ui.commit_table.item(row, 0).data(Qt.UserRole)
                excel_item = self.create_excel_item(row)

                if not excel_item.name:
                    self.__set_info("item is empty")
                    raise ValueError

                # Execute table to excel
                self.__set_info("Writing to Excel sheet...")
                write_to_excel(date, file, excel_item)
            except Exception as error:
                self.__set_info(f"Failed writing to excel sheet! Reason: {error}", Status.FAIL)
                self.logger.error(f"Failed on " f"{self.ui.commit_table.item(row, 1)}")
                self.logger.error(f"Error: {error}")
                return

        self.clean_table()
        self.logger.debug("Finished writing")
        self.__set_info("All done writing!", status=Status.DONE)

    def create_excel_item(self, row):
        excel_item = ExcelItem()
        excel_item.name = self.ui.commit_table.item(row, 1).data(Qt.UserRole)
        excel_item.vendor = self.ui.commit_table.item(row, 2).data(Qt.UserRole)
        excel_item.brand = self.ui.commit_table.item(row, 3).data(Qt.UserRole)
        excel_item.quantity = self.ui.commit_table.item(row, 4).data(Qt.UserRole)
        excel_item.unit = self.ui.commit_table.item(row, 5).data(Qt.UserRole)
        excel_item.cost = self.ui.commit_table.item(row, 6).data(Qt.UserRole)
        excel_item.isi = self.ui.commit_table.item(row, 8).data(Qt.UserRole)
        excel_item.isi_unit = self.ui.commit_table.item(row, 9).data(Qt.UserRole)
        excel_item.category = self.ui.commit_table.item(row, 11).data(Qt.UserRole)
        return excel_item

    def clean_table(self):
        self.ui.commit_table.clearContents()
        for row in reversed(range(self.ui.commit_table.rowCount())):
            self.ui.commit_table.removeRow(row)

    def __set_info(self, message, status: Status = Status.DEFAULT):
        """Display the info on the GUI

        done: green,
        fail: red,
        default: blue

        :param message: message to display on the gui
        :type message: str
        :param status: status of the message which effects the color of text
        :type status: str
        """
        color = status.value

        self.ui.status_bar.setText(message)
        self.ui.status_bar.setStyleSheet("color: {}".format(color))
        # noinspection PyUnresolvedReferences
        qApp.processEvents()  # update the UI


if __name__ == "__main__":
    app = QApplication(sys.argv)

    window = PembelianWidget(test=False)
    window.show()

    sys.exit(app.exec())
