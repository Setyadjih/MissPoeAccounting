from logging import getLogger

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font

from core.constants import ExcelItem, DATE_FORMAT, COMMA_FORMAT, RP_FORMAT, ITEM_INPUT_FORMAT, LOGGER_NAME
from core.utils import get_skip_list


def create_data_sheet(file, wb: Workbook, vendor_sheets):
    """Create named range of all vendors to count from. Overwrites any previous data and named range if sheet exists."""
    try:
        data_sheet: Worksheet = wb["DATA"]
    except KeyError:
        wb.create_sheet("DATA")
        data_sheet = wb["DATA"]
    for row, vendor in enumerate(vendor_sheets):
        data_sheet[f"A{row+1}"] = vendor
    new_range = DefinedName("Vendors", attr_text=f'DATA!$A$1:$A${len(vendor_sheets)}')
    # Delete and add new named range
    wb.defined_names.delete("Vendors")
    wb.defined_names.append(new_range)
    wb.save(file)


def clean_item_names(vendor_sheet: Worksheet):
    """Some entries have extra whitespace. This can mess with the cat entries, so we need to strip the names."""
    logger = getLogger(LOGGER_NAME)
    logger.debug("Cleaning sheet names.")
    for row in range(1, vendor_sheet.max_row + 1):
        item_name = vendor_sheet[f"B{row}"].value
        if not item_name:
            continue

        vendor_sheet[f"B{row}"] = str(item_name).strip()

    logger.debug("Finished Cleaning sheet names.")
    return vendor_sheet


def create_category_sheet(workbook, category):
    logger = getLogger(LOGGER_NAME)
    logger.info(f"Creating {category} in workbook")

    workbook.create_sheet(category)
    new_sheet: Worksheet = workbook[category]

    new_sheet["A1"] = "ITEM"
    new_sheet["A1"].font = Font(bold=True)
    new_sheet.merge_cells("A1:A2")

    new_sheet["B1"] = "UNIT BELI"
    new_sheet["B1"].font = Font(bold=True)
    new_sheet.merge_cells("B1:B2")

    new_sheet["C1"] = "UNIT ISI"
    new_sheet["C1"].font = Font(bold=True)
    new_sheet.merge_cells("C1:C2")

    new_sheet["D1"] = "MOV AVER"
    new_sheet["D1"].font = Font(bold=True)
    new_sheet["D2"] = "PRICE/UNIT"
    new_sheet["D2"].font = Font(bold=True)


def init_catsheet(file, categories: dict):
    """Clear out category sheets and recreate the entries"""
    logger = getLogger(LOGGER_NAME)
    # Due to openpyxl's structure, we need the data_only=False wb to save formula
    input_wb = openpyxl.load_workbook(file, data_only=False)
    clean_category_sheets(categories, input_wb)

    # default dict
    done_set = {"None", " "}
    skip_list = get_skip_list()

    # Iterate over all vendor sheets
    vendor_sheets = [_ for _ in input_wb.sheetnames if _ not in skip_list]
    create_data_sheet(file, input_wb, vendor_sheets)
    logger.debug(f"VENDORS: {vendor_sheets}")
    for sheet_name in vendor_sheets:
        if not sheet_name:
            continue

        sheet: Worksheet = input_wb[sheet_name]
        logger.info(f"Sheet: {sheet}")

        clean_sheet = clean_item_names(sheet)
        vendor_items = next(clean_sheet.iter_cols(min_col=2, max_col=2, min_row=3, values_only=True))
        for item in vendor_items:
            if not item or item in done_set:
                if item:
                    logger.debug(f"{item} is done, skipping")
                continue

            logger.debug(f"Look for '{item}' in {sheet_name}")
            row = vendor_items.index(item) + 3
            logger.debug(f"ROW: {row}, ITEM: {item}")
            excel_item = row_to_excel_item(clean_sheet, row)

            # Check if item is already in cat sheet
            category_items = next(input_wb[excel_item.category].iter_cols(1, 1, values_only=True))
            if item in category_items:
                logger.info("Item already in category, skipping")
                continue

            logger.info(f"Appending {excel_item.name} to {excel_item.category}")
            update_cat_avg(excel_item, input_wb)
            done_set.add(item)
    input_wb.save(file)
    logger.info("All done with init")


def row_to_excel_item(sheet: Worksheet, row):
    """Try to get data from row with clean up. If missing data, give defaults"""
    logger = getLogger(LOGGER_NAME)

    item_name = sheet[f"B{row}"].value
    sheet_title = sheet.title

    # Guard against missing units
    try:
        unit_isi = sheet[f"I{row}"].value
        if not unit_isi:
            unit_isi = sheet[f"E{row}"].value
    except IndexError:
        logger.error("Isi unit error, assigning g")
        unit_isi = "g"

    try:
        unit_beli = sheet[f"E{row}"].value
    except IndexError:
        logger.error("Beli unit error, assigning g")
        unit_beli = "g"
    # Check for common typos in categories
    try:
        category_value: str = sheet[f"K{row}"].value
    except IndexError:
        logger.error("Error getting Category, assigning Fresh")
        category_value = "Fresh"

    logger.debug(f"Category: {category_value}")
    return ExcelItem(name=item_name, vendor=sheet_title, unit_isi=unit_isi, unit_beli=unit_beli, category=category_value)


def clean_category_sheets(category_dict, input_wb):
    logger = getLogger(LOGGER_NAME)

    # Clear out all category sheets, leaving the header only
    for category in category_dict["CATEGORIES"]:
        # Create all missing category sheets
        try:
            category_sheet = input_wb[category]
        except KeyError:
            logger.warning(f"{category} not in workbook!")
            create_category_sheet(input_wb, category)
            category_sheet = input_wb[category]

        max_row = max(category_sheet.max_row, 3)
        logger.debug(f"Clearing {category} from 3 to {max_row}")
        category_sheet.delete_rows(3, max_row)


def write_to_excel(date, file, excel_item: ExcelItem):
    """Write the given data to the purchasing Excel sheet.

    :param str date: date of purchase
    :param str file: file path to Excel sheet to edit
    :param excel_item: ExcelItem with data
    :type excel_item: ExcelItem
    """
    logger = getLogger(LOGGER_NAME)

    # Load Excel file path
    # Need the data_only=False wb to save formula
    input_wb = openpyxl.load_workbook(file, data_only=False)

    # Create vendor sheet if new
    input_vendor = input_wb[excel_item.vendor]
    input_row = input_vendor.max_row + 1

    # iterate backwards until last_row is after a row with data
    while not input_vendor[f"B{input_row-1}"].value:
        input_row -= 1

    # Hard coded minimum to not clash with merged cells:
    if input_row < 3:
        input_row = 3

    logger.debug(f"max row = {input_vendor.max_row}, input row = {input_row}")
    logger.debug(f"Appending item data")

    # Generic columns
    input_vendor[f"A{input_row}"] = date
    input_vendor[f"G{input_row}"] = f"=D{input_row}*F{input_row}"
    input_vendor[f"J{input_row}"] = f"=G{input_row}/H{input_row}"

    # item specific data
    for column in ITEM_INPUT_FORMAT.keys():
        input_vendor[f"{column}{input_row}"] = getattr(excel_item, ITEM_INPUT_FORMAT[column])

    # format cells for Rupiah
    logger.debug("Assigning format strings")
    date_cell = input_vendor.cell(input_row, column_index_from_string("A"))
    date_cell.number_format = DATE_FORMAT

    cost_cell = input_vendor.cell(input_row, column_index_from_string("F"))
    cost_cell.number_format = RP_FORMAT

    total_cell = input_vendor.cell(input_row, column_index_from_string("G"))
    total_cell.number_format = RP_FORMAT

    isi_cell = input_vendor.cell(input_row, column_index_from_string("H"))
    isi_cell.number_format = COMMA_FORMAT

    per_unit_cell = input_vendor.cell(input_row, column_index_from_string("J"))
    per_unit_cell.number_format = RP_FORMAT

    logger.debug(f"Assigning {excel_item.name} to {excel_item.category}")
    update_cat_avg(excel_item, input_wb)

    input_wb.save(file)


def update_cat_avg(excel_item, workbook):
    """Calculate average price for each item, total quantity, total units.
    Assumes J is the price/unit column, and B is the name column
    :param excel_item: ExcelItem with data
    :param workbook: workbook to read from
    """
    logger = getLogger(LOGGER_NAME)

    category = excel_item.category
    logger.debug(f"Checking Item: {excel_item.name} in Category: {excel_item.category}.")

    # check if item in list
    cat_items = next(workbook[category].iter_cols(1, 1, values_only=True))
    if excel_item.name not in cat_items:
        init_formula(excel_item, workbook)
    else:
        logger.debug(f"Item: {excel_item.name} already entered.")


def get_avg_price_formula(row):
    return f"""=SUMPRODUCT(SUMIF(INDIRECT("'"&Vendors&"'!"&"B:B"),A{row}, INDIRECT("'"&Vendors&"'!"&"J:J"))) \
/ SUMPRODUCT(COUNTIF(INDIRECT("'"&Vendors&"'!"&"B:B"), A{row}))
"""


def get_max_price_formula(row):
    return f"""=MAX(MAXIFS(INDIRECT("'"&Vendors&"'!"&"J:J"), INDIRECT("'"&Vendors&"'!"&"B:B"), A{row}))"""


def init_formula(excel_item: ExcelItem, workbook, row=None):
    """Generate full average formula. Get the row as a check in the book."""
    logger = getLogger(LOGGER_NAME)
    logger.info("Item is new, adding and initializing formula")

    category = excel_item.category
    ws: Worksheet = workbook[category]

    # Set row to given, max_row, or minimum 3
    if not row:
        logger.debug(f"Creating {category} entry")
        workbook[category].append({"A": excel_item.name, "B": excel_item.unit_beli, "C": excel_item.unit_isi})
        row = workbook[category].max_row
    row = row if row > 3 else 3

    ws[f"D{row}"] = get_avg_price_formula(row)
    ws[f"D{row}"].number_format = RP_FORMAT

    ws[f"E{row}"] = get_max_price_formula(row)
    ws.formula_attributes[f"E{row}"] = {"t": "array", "ref": f"E{row}:E{row}"}
    ws[f"E{row}"].number_format = RP_FORMAT


def transfer_records(old_workbook_path, new_workbook_path, categories: dict):
    """Check entries from old to new, append any missing to new"""
    logger = getLogger(LOGGER_NAME)

    old_workbook = openpyxl.load_workbook(old_workbook_path, data_only=True)
    new_workbook = openpyxl.load_workbook(new_workbook_path, data_only=True)
    new_workbook_input = openpyxl.load_workbook(new_workbook_path, data_only=False)
    try:
        # Use old wb name as a fake vendor
        item_category = new_workbook_input["_IMPORT_"]
    except KeyError:
        new_workbook_input.create_sheet("_IMPORT_")
        item_category = new_workbook_input["_IMPORT_"]

    # Get item entries as sets
    logger.debug("Generating item lists...")
    logger.debug("Getting items from old workbook...")
    old_cat_items = get_items_in_category(old_workbook, categories)
    logger.debug("Getting items from new workbook...")

    # Append old items to new workbook
    for item_name in old_cat_items.keys():
        logger.debug(f"Found old item: {str(item_name).strip()}")
        item_column_details = {
            "B": str(item_name).strip(),
            "E": old_cat_items[item_name]["unit_beli"],
            "I": old_cat_items[item_name]["unit_isi"],
            "J": old_cat_items[item_name]["unit_price"],
            "K": old_cat_items[item_name]["category"],
        }
        item_category.append(item_column_details)

    logger.debug("Saving and beginning init")
    new_workbook_input.save(new_workbook_path)
    init_catsheet(new_workbook_path, categories)
    logger.debug("Finished transfer!")


def get_items_in_category(workbook, categories):
    logger = getLogger(LOGGER_NAME)
    # Get new entries
    category_items = {}
    logger.debug("Checking info from workbook")
    for category in categories["CATEGORIES"]:
        try:
            category_sheet: Worksheet = workbook[category]
        except KeyError:
            logger.warning(f"Could not find Worksheet '{category}'. Will skip this.")
            continue

        for row in range(3, category_sheet.max_row):
            item = {
                "category": category,
                "name": category_sheet[f"A{row}"].value,
                "unit_beli": category_sheet[f"B{row}"].value,
                "unit_isi": category_sheet[f"C{row}"].value,
                "unit_price": category_sheet[f"D{row}"].value,
            }
            category_items[item["name"]] = item

    return category_items
