import sys
from datetime import datetime
import shutil
from pathlib import Path

from PySide2.QtWidgets import (
    QApplication,
    QWidget,
    QTableWidgetItem,
    QFileDialog,
    QAction,
    QMessageBox
)
from PySide2.QtCore import Qt
from openpyxl import load_workbook

from utils import get_logger
from resources.pembelian_ui_ss import Ui_pembelian
from excel_functions import write_to_excel, init_catsheet
import constants
from constants import CAT_REF


# noinspection SpellCheckingInspection
class PembelianWidget(QWidget):
    def __init__(self, parent=None):
        super(PembelianWidget, self).__init__(parent)
        self.ui = Ui_pembelian()
        self.ui.setupUi(self)
        self.ui.date_line.setText(constants.DATE)
        self.logger = get_logger("excel_automator")
        self.logger.info("Initializing program")

        # Context menu setup
        self.ui.commit_table.setContextMenuPolicy(Qt.ActionsContextMenu)
        self.del_row_action = QAction(self, text="Delete Table Row")
        # noinspection PyUnresolvedReferences
        self.del_row_action.triggered.connect(self.delete_table_row)
        self.ui.commit_table.addAction(self.del_row_action)

        self.setWindowTitle(f"Poe Excel Automator {constants.APP_VERSION}")
        self.ui.status_bar.setText("Ready for input.")
        self.ui.confirm_button.setDisabled(True)
        self.ui.init_button.setDisabled(True)

        # Default test button to hide
        self.ui.test_button.hide()

        # Setup Category list
        self.categories = {"MISC": [], "CATEGORIES": []}

        # category reference check
        if not Path(CAT_REF).exists():
            self.logger.info("Creating new cat ref file")
            with open(CAT_REF, "w") as new_file:
                new_file.write(constants.DEFAULT_CATEGORIES)
            message = QMessageBox()
            message.setWindowTitle("Default categories file created")
            message.setText(f"A default category list was created. Please "
                            f"edit the {CAT_REF} file and rerun "
                            f"the program if you need to add categories")
            message.exec_()

        with open(CAT_REF) as cat_ref:
            cat_flag = False
            for line in cat_ref.readlines():
                line = line.rstrip()
                if not line:
                    continue
                # Flag check
                if line == "[CATEGORIES]":
                    cat_flag = True
                    continue
                elif line == "[MISC]":
                    cat_flag = False
                    continue

                # input to category dictionary
                if cat_flag:
                    self.logger.debug(f"Adding category {line}")
                    self.categories["CATEGORIES"].append(line)
                else:
                    self.logger.debug(f"Adding misc {line}")
                    self.categories["MISC"].append(line)

        self.ui.category_combo.clear()
        self.ui.category_combo.addItems(self.categories["CATEGORIES"])

        # Hookup buttons
        self.ui.test_button.clicked.connect(self.test_func)

        self.ui.add_vendor_button.clicked.connect(self.add_to_table)
        self.ui.file_browse_button.clicked.connect(self.get_excel_sheet)
        self.ui.confirm_button.clicked.connect(self.confirm_table)
        self.ui.init_button.clicked.connect(self.init_cat_button)

        # # FIXME:
        # TESTING PARAMS
        # self.ui.xls_file_browser.setText(
        #     "D:\Miss Poe\Costings\_data\Pembelian 2020_TESTING_CLEANED.xlsx"
        # )
        # self.ui.test_button.show()

        self.ui.confirm_button.setDisabled(True)
        self.ui.init_button.setDisabled(True)

    def test_func(self):
        purchase_book = load_workbook(self.ui.xls_file_browser.text())
        for vendor in purchase_book.sheetnames:
            self.ui.vendor_combo.addItem(vendor)

    def make_backup(self):
        # Create a backup copy just in case
        self.logger.info("Saving backup.")
        file = self.ui.xls_file_browser.text()
        dest = Path(file).with_suffix(".bak")
        shutil.copyfile(file, dest)

    def init_cat_button(self):
        file = self.ui.xls_file_browser.text()

        self.make_backup()

        # Setting up warning box to be sure
        warning = QMessageBox()
        warning.setIcon(QMessageBox.Warning)
        warning.setText(
            "Initializing the category data can take a long time, are you sure"
            " you want to do this?\n Please make sure to clear out the "
            "category sheets before hand."
        )
        warning.setWindowTitle("Are you sure?")

        # I don't really understand the pipe operator, but it works
        warning.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)

        # result is an int, compare to qmessage status works
        result = warning.exec_()
        if result == QMessageBox.Cancel:
            return

        self.__set_info("Working on data....")
        try:
            init_catsheet(file, self.categories, self.logger)
        except Exception as e:
            self.logger.error(e)
            self.__set_info(f"Failed to init data! Error: {e}", "fail")
            return
        self.__set_info("All done!", "done")

    def delete_table_row(self):
        current_row = self.ui.commit_table.currentRow()
        self.ui.commit_table.removeRow(current_row)

    def get_excel_sheet(self):
        """Load Purchase Excelsheet and get vendors"""
        try:
            self.ui.xls_file_browser.setText(
                QFileDialog.getOpenFileName(filter="Excel sheets (*.xlsx)")[0]
            )
        except KeyError as error:
            self.__set_info(f"Failed to pick sheet! Vendor doesn't exist.", "fail")
            self.logger.error(error)
            return
        except Exception as error:
            self.__set_info(f"Failed to pick sheet! Reason: {error}", "fail")
            self.logger.error(error)
            return

        if not self.ui.xls_file_browser.text():
            return
        purchase_book = load_workbook(self.ui.xls_file_browser.text())
        for vendor in purchase_book.sheetnames:
            self.ui.vendor_combo.addItem(vendor)

        self.ui.confirm_button.setDisabled(True)
        self.ui.init_button.setDisabled(True)

        if self.ui.xls_file_browser.text():
            self.ui.confirm_button.setEnabled(True)
            self.ui.init_button.setEnabled(True)

    def clear_inputs(self):
        self.ui.vendor_combo.clear()
        self.ui.item_line.clear()
        self.ui.qty_spin.clear()
        self.ui.unit_line.clear()
        self.ui.harga_spin.clear()
        self.ui.isi_spin.clear()
        self.ui.isi_unit_line.clear()
        self.__set_info("Cleared inputs!", status="done")

    def add_to_table(self):
        # Table entry validation
        if (
            self.ui.harga_spin.value() == 0
            or self.ui.isi_spin.value() == 0
            or self.ui.qty_spin.value() == 0
            or not self.ui.date_line.text()
            or not self.ui.isi_unit_combo.currentText()
            or not self.ui.item_line.text()
            or not self.ui.vendor_combo.currentText()
        ):
            self.__set_info("Values cannot be 0!", "fail")
            return

        date_text = self.ui.date_line.text()
        date = datetime.strptime(date_text, "%d/%m/%y")

        # Commit input to table
        row_count = self.ui.commit_table.rowCount()
        self.ui.commit_table.insertRow(row_count)
        new_row = row_count

        total_cost = self.ui.qty_spin.value() * self.ui.harga_spin.value()
        unit_cost = total_cost / self.ui.isi_spin.value()

        date_data = QTableWidgetItem(self.ui.date_line.text())
        date_data.setData(Qt.UserRole, date)

        vendor_data = QTableWidgetItem(self.ui.vendor_combo.currentText())
        vendor_data.setData(Qt.UserRole, self.ui.vendor_combo.currentText())

        if not self.ui.merek_line.text():
            self.ui.merek_line.setText("")
        merek_data = QTableWidgetItem(self.ui.merek_line.text())
        merek_data.setData(Qt.UserRole, self.ui.merek_line.text())

        item_data = QTableWidgetItem(self.ui.item_line.text())
        item_data.setData(Qt.UserRole, self.ui.item_line.text())

        qty_data = QTableWidgetItem(self.ui.qty_spin.text())
        qty_data.setData(Qt.UserRole, self.ui.qty_spin.value())

        unit_data = QTableWidgetItem(self.ui.unit_combo.currentText())
        unit_data.setData(Qt.UserRole, self.ui.unit_combo.currentText())

        harga_data = QTableWidgetItem(self.ui.harga_spin.text())
        harga_data.setData(Qt.UserRole, self.ui.harga_spin.value())

        total_data = QTableWidgetItem(str(total_cost))
        total_data.setData(Qt.UserRole, total_cost)

        isi_data = QTableWidgetItem(self.ui.isi_spin.text())
        isi_data.setData(Qt.UserRole, self.ui.isi_spin.value())

        isi_unit_data = QTableWidgetItem(self.ui.isi_unit_combo.currentText())
        isi_unit_data.setData(Qt.UserRole, self.ui.isi_unit_combo.currentText())

        unit_harga_data = QTableWidgetItem(str(unit_cost))
        unit_harga_data.setData(Qt.UserRole, unit_cost)

        category_data = QTableWidgetItem(self.ui.category_combo.currentText())
        category_data.setData(Qt.UserRole, self.ui.category_combo.currentText())

        details = (
            date_data,
            item_data,
            vendor_data,
            merek_data,
            qty_data,
            unit_data,
            harga_data,
            total_data,
            isi_data,
            isi_unit_data,
            unit_harga_data,
            category_data,
        )

        for column, item in enumerate(details):
            self.ui.commit_table.setItem(new_row, column, item)
        self.__set_info("Adding item to table...")

    def confirm_table(self):
        self.make_backup()
        self.logger.info("Executing table")
        file = self.ui.xls_file_browser.text()
        if self.ui.commit_table.rowCount() == 0:
            self.__set_info("Nothing to write")
            return

        for row in range(self.ui.commit_table.rowCount()):
            try:
                # Get values from item ranges
                date = self.ui.commit_table.item(row, 0).data(Qt.UserRole)
                item = self.ui.commit_table.item(row, 1).data(Qt.UserRole)
                vendor = self.ui.commit_table.item(row, 2).data(Qt.UserRole)
                merek = self.ui.commit_table.item(row, 3).data(Qt.UserRole)
                quantity = self.ui.commit_table.item(row, 4).data(Qt.UserRole)
                unit = self.ui.commit_table.item(row, 5).data(Qt.UserRole)
                harga = self.ui.commit_table.item(row, 6).data(Qt.UserRole)
                isi = self.ui.commit_table.item(row, 8).data(Qt.UserRole)
                isi_unit = self.ui.commit_table.item(row, 9).data(Qt.UserRole)
                category = self.ui.commit_table.item(row, 11).data(Qt.UserRole)

                if not item:
                    self.__set_info("item is empty")
                    return

                # Execute table to excel
                write_to_excel(
                    self.categories,
                    date,
                    file,
                    vendor,
                    merek,
                    item,
                    quantity,
                    unit,
                    harga,
                    isi,
                    isi_unit,
                    category,
                    self.logger,
                )

                self.__set_info("Writing to Excel sheet...")
            except Exception as error:
                self.logger.error(f"Failed on "
                                  f"{self.ui.commit_table.item(row, 1)}")
                self.__set_info(f"Failed writing to excel sheet! Reason: {error}", "fail")
                self.logger.error(f"Error: {error}")
                return
        self.clean_table()

        self.__set_info("All done writing!", status="done")

    def clean_table(self):
        self.ui.commit_table.clearContents()
        for row in reversed(range(self.ui.commit_table.rowCount())):
            self.ui.commit_table.removeRow(row)

    def __set_info(self, message, status=""):
        """Display the info on the GUI

        :param message: message to display on the gui
        :type message: str
        :param status: status of the message which effects the color of text
        :type status: str
        """
        if status == "fail":
            color = "red"
        elif status == "done":
            color = "#00ff06"  # bright green
        else:
            color = "#00b2ff"  # bright blue

        self.ui.status_bar.setText(message)
        self.ui.status_bar.setStyleSheet("color: {}".format(color))
        # noinspection PyUnresolvedReferences
        qApp.processEvents()  # update the UI


if __name__ == "__main__":
    app = QApplication(sys.argv)

    window = PembelianWidget()
    window.show()

    sys.exit(app.exec_())
